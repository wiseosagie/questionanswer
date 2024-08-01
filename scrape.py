import streamlit as st
import requests
import pandas as pd
import re
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse
from collections import deque, defaultdict
import logging
import time
import os
from openpyxl import load_workbook, Workbook
from dotenv import load_dotenv
from datetime import datetime

# Load API keys from .env file
load_dotenv()
GOOGLE_PLACES_API_KEY = os.getenv("GOOGLE_PLACES_API_KEY")
GOOGLE_GEOCODING_API_KEY = os.getenv("GOOGLE_GEOCODING_API_KEY")

# Set up logging
logging.basicConfig(level=logging.INFO)

GOOGLE_PLACES_API_URL = "https://maps.googleapis.com/maps/api/place/textsearch/json"
GOOGLE_PLACE_DETAILS_API_URL = "https://maps.googleapis.com/maps/api/place/details/json"
GOOGLE_GEOCODING_API_URL = "https://maps.googleapis.com/maps/api/geocode/json"

# Estimated cost per API call (in USD)
GEOCODING_API_COST = 0.005  # Example cost for a geocoding API call
PLACES_API_COST = 0.017  # Example cost for a Places API call

# Estimated time per API call (in seconds)
GEOCODING_API_TIME = 0.5  # Example time for a geocoding API call
PLACES_API_TIME = 0.5  # Example time for a Places API call
SCRAPING_TIME = 2  # Example time for scraping each website

def generate_grid_coordinates(lat, lon, area):
    if area < 100:
        grid_cell_size = 1  # miles
    elif 100 <= area < 1000:
        grid_cell_size = 3  # miles
    elif 1000 <= area < 10000:
        grid_cell_size = 5  # miles
    elif 10000 <= area < 50000:
        grid_cell_size = 15  # miles
    else:
        grid_cell_size = 25  # miles

    # Convert grid cell size to degrees
    grid_cell_lat = grid_cell_size / 69.0  # 1 degree latitude ~ 69 miles
    grid_cell_lon = grid_cell_size / 55.0  # 1 degree longitude ~ 55 miles at mid-latitudes

    # Estimate bounds (these should be replaced with real bounds if available)
    north = lat + (area**0.5 / 69.0) / 2
    south = lat - (area**0.5 / 69.0) / 2
    east = lon + (area**0.5 / 55.0) / 2
    west = lon - (area**0.5 / 55.0) / 2

    coordinates = []
    current_lat = south
    while current_lat <= north:
        current_lon = west
        while current_lon <= east:
            coordinates.append((current_lat, current_lon))
            current_lon += grid_cell_lon
        current_lat += grid_cell_lat
    
    return coordinates

def get_coordinates(location, area):
    params = {
        'address': location,
        'key': GOOGLE_GEOCODING_API_KEY
    }
    try:
        response = requests.get(GOOGLE_GEOCODING_API_URL, params=params)
        response.raise_for_status()
        location_data = response.json().get('results', [])[0]['geometry']['location']
        lat = location_data['lat']
        lon = location_data['lng']
        return generate_grid_coordinates(lat, lon, area), len(generate_grid_coordinates(lat, lon, area))
    except requests.RequestException as e:
        logging.error(f"Could not geocode location: {location} - {e}")
        return [], 0

def extract_emails(text):
    # Regex to find lines containing '@' and '.com'
    email_pattern = r'[\w\.-]+@[\w\.-]+\.\w+'
    return re.findall(email_pattern, text)

def extract_emails_from_tags(soup):
    keywords = ["email", "emails", "contact", "contact us"]
    emails = set()
    for tag in soup.find_all(True):
        if any(keyword in tag.get_text().lower() for keyword in keywords):
            text_content = tag.get_text()
            emails.update(extract_emails(text_content))
    return emails

def extract_phone_numbers(text):
    phone_patterns = [
        r'(?:(?:\+?(\d{1,3}))?[-.\s]?)?(\(?\d{3}\)?[-.\s]?)?\d{3}[-.\s]?\d{4,6}',
        r'\b(?:Tel|TEL|Telephone|phone|contact)[:\s]*((?:\(?\d{3}\)?\s*-?\s*\d{3}\s*-?\s*\d{4,5}))\b'
    ]
    phone_numbers = []
    for pattern in phone_patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        for match in matches:
            if isinstance(match, tuple):
                match = ''.join(match)  # Join tuple elements into a single string
            phone_number = re.sub(r'\D', '', match)  # Remove all non-digit characters
            if len(phone_number) == 10:  # Ensure the phone number is exactly 10 digits long
                phone_numbers.append(phone_number)
    return phone_numbers

def is_valid_url(url, base_url):
    parsed_url = urlparse(url)
    return parsed_url.scheme in {"http", "https"} and parsed_url.netloc == urlparse(base_url).netloc

def verify_url(url):
    try:
        response = requests.head(url, allow_redirects=True, timeout=5)
        response.raise_for_status()
        return True
    except requests.RequestException as e:
        logging.error(f"URL verification failed for {url}: {e}")
        return False

def web_crawler(start_url, total_urls, status_placeholder, stop_signal):
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }
    visited_urls = set()
    urls_to_visit = deque([start_url])
    all_emails = set()
    all_phone_numbers = set()
    max_retries = 3

    url_404_count = defaultdict(int)
    domain_404_count = defaultdict(int)
    urls_scraped = 0

    while urls_to_visit and not stop_signal.get('stop', False) and urls_scraped < 100:
        url = urls_to_visit.popleft()
        domain = urlparse(url).netloc

        if url in visited_urls or domain_404_count[domain] > 20:
            continue

        retry_count = 0
        while retry_count < max_retries:
            try:
                response = requests.get(url, headers=headers, timeout=10)
                response.raise_for_status()
                
                # Check if the content type is HTML
                if 'text/html' in response.headers.get('Content-Type', ''):
                    visited_urls.add(url)
                    page_content = response.text
                    try:
                        soup = BeautifulSoup(page_content, 'html.parser')
                    except Exception as e:
                        logging.error(f"Failed to parse HTML content for {url}: {e}")
                        break  # Exit retry loop if parsing fails

                    text_content = soup.get_text()
                    all_emails.update(extract_emails(text_content))
                    all_emails.update(extract_emails_from_tags(soup))
                    all_phone_numbers.update(extract_phone_numbers(text_content))
                    
                    for link in soup.find_all('a', href=True):
                        link_url = urljoin(start_url, link['href'])
                        if is_valid_url(link_url, start_url) and link_url not in visited_urls:
                            urls_to_visit.append(link_url)
                    
                    urls_scraped += 1
                    status_placeholder.text(f"Scraped {urls_scraped} of {total_urls} pages. Emails found: {len(all_emails)}")

                    break  # Exit retry loop if successful
                else:
                    logging.error(f"Non-HTML content at {url}, skipping.")
                    break
            except requests.HTTPError as e:
                if response.status_code == 404:
                    logging.error(f"404 Client Error: Not Found for url: {url}")
                    url_404_count[url] += 1
                    domain_404_count[domain] += 1

                    if url_404_count[url] > 10:
                        logging.error(f"Skipping URL {url} due to excessive 404 errors.")
                        break  # Skip this URL

                    if domain_404_count[domain] > 20:
                        logging.error(f"Skipping domain {domain} due to excessive 404 errors.")
                        break  # Skip this domain
                else:
                    logging.error(f"HTTPError for {url}: {e}")
                    break  # Exit retry loop for other HTTP errors
            except (requests.ConnectionError, requests.Timeout) as e:
                logging.error(f"Failed to retrieve the webpage {url}: {e}")
                retry_count += 1
                time.sleep(2 ** retry_count)  # Exponential backoff
            except requests.RequestException as e:
                logging.error(f"RequestException for {url}: {e}")
                break

    return all_emails, all_phone_numbers

def google_maps_search(query, lat, lon):
    params = {
        'query': query,
        'location': f'{lat},{lon}',
        'radius': 5000,
        'key': GOOGLE_PLACES_API_KEY
    }
    try:
        response = requests.get(GOOGLE_PLACES_API_URL, params=params)
        response.raise_for_status()
        results = response.json().get('results', [])
        return results[:10]  # Limit to 10 businesses
    except requests.RequestException as e:
        logging.error(f"Failed to search Google Maps: {e}")
        return []

def get_place_details(place_id):
    params = {
        'place_id': place_id,
        'key': GOOGLE_PLACES_API_KEY
    }
    try:
        response = requests.get(GOOGLE_PLACE_DETAILS_API_URL, params=params)
        response.raise_for_status()
        return response.json().get('result', {})
    except requests.RequestException as e:
        logging.error(f"Failed to retrieve place details: {e}")
        return {}

def update_admin_record(data, query, location, total_cost):
    admin_record_path = os.path.join('static', 'admin_record.xlsx')
    if os.path.exists(admin_record_path):
        wb = load_workbook(admin_record_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)  # Remove the default sheet

    sheet_name = f"{query}_{location}".replace(" ", "_")
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(['Business Name', 'Website', 'Phone Numbers', 'Emails', 'Place ID', 'Query', 'Location', 'Total Cost'])

    for row in data:
        ws.append([
            row['Business Name'] or "N/A",
            row['Website'] or "N/A",
            row['Phone Numbers'] or "N/A",
            row['Emails'] or "N/A",
            row['Place ID'] or "N/A",
            query,
            location,
            total_cost
        ])

    wb.save(admin_record_path)

def log_search_start(query, location, area, num_businesses):
    log_file_path = 'static/search_log.csv'
    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    search_data = {
        'Query': query,
        'Location': location,
        'Area': area,
        'Number of Businesses': num_businesses,
        'Event': 'Start',
        'Date': current_time
    }
    log_df = pd.DataFrame([search_data])
    
    if os.path.exists(log_file_path):
        log_df.to_csv(log_file_path, mode='a', header=False, index=False)
    else:
        log_df.to_csv(log_file_path, mode='w', header=True, index=False)

def log_search_end(query, location, area, num_businesses, num_emails, num_phone_numbers, total_cost, total_time, interrupted=False):
    log_file_path = 'static/search_log.csv'
    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    search_data = {
        'Query': query,
        'Location': location,
        'Area': area,
        'Number of Businesses': num_businesses,
        'Number of Emails': num_emails,
        'Number of Phone Numbers': num_phone_numbers,
        'Total Cost': total_cost,
        'Total Time': total_time,
        'Event': 'Interrupted' if interrupted else 'End',
        'Date': current_time
    }
    log_df = pd.DataFrame([search_data])
    
    if os.path.exists(log_file_path):
        log_df.to_csv(log_file_path, mode='a', header=False, index=False)
    else:
        log_df.to_csv(log_file_path, mode='w', header=True, index=False)

# Streamlit app
def main():
    st.title("Business Data Extractor")

    query = st.text_input("Enter the business type (e.g., restaurant, salon)")
    location = st.text_input("Enter the location (e.g., city, address)")
    area = st.number_input("Enter the search area in square miles", min_value=1, max_value=100000, value=100)
    num_businesses = st.number_input("Enter the number of businesses to find", min_value=1, max_value=100, value=10)

    # Session state to handle stop signal and scraping progress
    if "stop_signal" not in st.session_state:
        st.session_state.stop_signal = {"stop": False}
    if "scraping_completed" not in st.session_state:
        st.session_state.scraping_completed = False
    if "scraped_data" not in st.session_state:
        st.session_state.scraped_data = []
    if "total_cost" not in st.session_state:
        st.session_state.total_cost = 0
    if "total_time" not in st.session_state:
        st.session_state.total_time = 0
    if "total_emails" not in st.session_state:
        st.session_state.total_emails = set()
    if "total_phone_numbers" not in st.session_state:
        st.session_state.total_phone_numbers = set()

    if st.button("Search") and not st.session_state.scraping_completed:
        st.session_state.stop_signal["stop"] = False
        if query and location:
            with st.spinner("Generating coordinates..."):
                coordinates, num_coords = get_coordinates(location, area=area)
                if not coordinates:
                    st.error("Could not generate coordinates. Please check the location.")
                    return

            log_search_start(query, location, area, num_businesses)

            status_placeholder = st.empty()
            stop_button_placeholder = st.empty()

            with st.spinner("Searching..."):
                while not st.session_state.scraping_completed:
                    stop_button = stop_button_placeholder.button("Stop Search")
                    if stop_button:
                        st.session_state.stop_signal["stop"] = True

                    for lat, lon in coordinates:
                        st.write(f"Searching at coordinates: ({lat}, {lon})")
                        search_results = google_maps_search(query, lat, lon)
                        st.session_state.total_cost += PLACES_API_COST  # Cost for each Places API search
                        st.session_state.total_time += PLACES_API_TIME  # Time for each Places API search
                        
                        for result in search_results:
                            place_id = result['place_id']
                            details = get_place_details(place_id)
                            business_name = details.get('name')
                            website = details.get('website')
                            phone_number = details.get('formatted_phone_number')
                            
                            if website and verify_url(website):
                                total_urls = len(coordinates) * 10
                                emails, phone_numbers = web_crawler(website, total_urls, status_placeholder, st.session_state.stop_signal)
                                st.session_state.total_time += SCRAPING_TIME  # Time for scraping each website
                                st.session_state.total_emails.update(emails)
                                st.session_state.total_phone_numbers.update(phone_numbers)
                            else:
                                emails, phone_numbers = set(), set()
                            
                            st.session_state.scraped_data.append({
                                'Business Name': business_name or "N/A",
                                'Website': website or "N/A",
                                'Phone Numbers': phone_number or "N/A",
                                'Emails': ', '.join(emails) or "N/A",
                                'Place ID': place_id or "N/A"
                            })

                            if len(st.session_state.scraped_data) >= num_businesses or st.session_state.stop_signal["stop"]:
                                break

                        if len(st.session_state.scraped_data) >= num_businesses or st.session_state.stop_signal["stop"]:
                            break

                    st.session_state.scraping_completed = True

                # Update admin record and log search details regardless of interruption
                update_admin_record(st.session_state.scraped_data, query, location, st.session_state.total_cost)
                log_search_end(query, location, area, len(st.session_state.scraped_data), len(st.session_state.total_emails), len(st.session_state.total_phone_numbers), st.session_state.total_cost, st.session_state.total_time, interrupted=st.session_state.stop_signal["stop"])

            stop_button_placeholder.empty()  # Remove the stop button
            total_cost_multiplied = st.session_state.total_cost * 5  # Multiply the total cost by 5 for display
            st.write(f"Total number of results found: {len(st.session_state.scraped_data)}")
            st.write(f"Estimated cost of the search: ${total_cost_multiplied:.2f}")  # Display the multiplied cost
            st.write(f"Estimated time for the search: {st.session_state.total_time:.2f} seconds")

            df = pd.DataFrame(st.session_state.scraped_data[:num_businesses]).drop_duplicates()  # Ensure no more than requested results
            df.fillna("N/A", inplace=True)  # Fill empty fields with "N/A"
            file_path = 'business_data.xlsx'
            df.to_excel(file_path, index=False)

            st.success("Search completed!")
            st.download_button("Download Business Data", data=open(file_path, 'rb').read(), file_name='business_data.xlsx')

    if st.session_state.scraping_completed and st.button("Continue Scraping"):
        st.session_state.scraping_completed = False
        st.experimental_rerun()

    if st.session_state.scraping_completed and st.button("Stop and Save Data"):
        # Save the data and clear the session state
        total_cost_multiplied = st.session_state.total_cost * 5  # Multiply the total cost by 5 for display
        st.write(f"Total number of results found: {len(st.session_state.scraped_data)}")
        st.write(f"Estimated cost of the search: ${total_cost_multiplied:.2f}")  # Display the multiplied cost
        st.write(f"Estimated time for the search: {st.session_state.total_time:.2f} seconds")

        df = pd.DataFrame(st.session_state.scraped_data).drop_duplicates()  # Ensure no more than requested results
        df.fillna("N/A", inplace=True)  # Fill empty fields with "N/A"
        file_path = 'business_data.xlsx'
        df.to_excel(file_path, index=False)

        st.download_button("Download Business Data", data=open(file_path, 'rb').read(), file_name='business_data.xlsx')

if __name__ == "__main__":
    if not os.path.exists('static'):
        os.makedirs('static')
    main()
